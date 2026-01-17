"""Excel export service using openpyxl."""

from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.action_plan import ActionPlan
from ..config.settings import settings


class ExcelService:
    """Service for exporting data to Excel spreadsheets."""
    
    def export_action_plan(self, action_plan: ActionPlan, output_path: str) -> bool:
        """Export action plan to Excel.
        
        Args:
            action_plan: ActionPlan instance to export.
            output_path: Path where Excel file will be saved.
        
        Returns:
            True if successful, False otherwise.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Action Plan"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Task Description',
                'Responsible Person',
                'Deadline',
                'Status',
                'Priority',
                'Notes',
                'Completed Date'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Set column widths
            column_widths = [40, 20, 15, 15, 12, 30, 15]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # Data rows
            for row_num, task in enumerate(action_plan.tasks, 2):
                data = [
                    task.task_description,
                    task.responsible_person or '',
                    task.deadline.strftime('%Y-%m-%d') if task.deadline else '',
                    task.status.value,
                    task.priority.value,
                    task.notes or '',
                    task.completed_date.strftime('%Y-%m-%d') if task.completed_date else ''
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = border
                    
                    # Color code by status
                    if col_num == 4:  # Status column
                        if value == "Completed":
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        elif value == "In Progress":
                            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        elif value == "Blocked":
                            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    
                    # Color code by priority
                    if col_num == 5:  # Priority column
                        if value == "Critical":
                            cell.font = Font(color="E81123", bold=True)
                        elif value == "High":
                            cell.font = Font(color="FF8C00", bold=True)
            
            # Add summary section
            summary_row = len(action_plan.tasks) + 3
            
            ws.cell(row=summary_row, column=1).value = "Summary:"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
            
            total_tasks = len(action_plan.tasks)
            completed_tasks = sum(1 for task in action_plan.tasks if task.status.value == "Completed")
            completion_pct = action_plan.calculate_completion()
            
            ws.cell(row=summary_row + 1, column=1).value = "Total Tasks:"
            ws.cell(row=summary_row + 1, column=2).value = total_tasks
            
            ws.cell(row=summary_row + 2, column=1).value = "Completed Tasks:"
            ws.cell(row=summary_row + 2, column=2).value = completed_tasks
            
            ws.cell(row=summary_row + 3, column=1).value = "Completion %:"
            ws.cell(row=summary_row + 3, column=2).value = f"{completion_pct}%"
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
